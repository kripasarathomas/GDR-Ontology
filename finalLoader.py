import pandas as pd
from openpyxl import load_workbook
import warnings
from pathlib import Path

warnings.simplefilter("ignore", UserWarning)


def load_GDR_data_from_file(filepath: Path):
    """Loads GDR classification data from all sheets in a single Excel file."""
    try:
        filepath = str(filepath)  # Convert Path object to string

        if not filepath.endswith(".xlsx"):
            raise ValueError(f"Unsupported file format: {filepath}. Use an Excel (.xlsx) file.")

        wb = load_workbook(filename=filepath, data_only=True)
        benennungen = {}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"Processing sheet: {sheet_name} in file {filepath}")

            # Extract job titles (Column A) and classification codes (Column B)
            bezlabel = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]
            bezlabelkey = [sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row + 1)]

            # Aggregate into dictionary
            for job_title, code in zip(bezlabel, bezlabelkey):
                if code:
                    benennungen.setdefault(code, []).append(job_title)

        return benennungen

    except Exception as e:
        print(f"Error loading {filepath}: {e}")
        return {}


def load_GDR_data_from_folder(folder_path: Path):
    """Loads GDR classification data from all Excel files in a folder."""
    all_data = {}
    folder_path = Path(folder_path)
    if not folder_path.exists() or not folder_path.is_dir():
        print(f" Folder not found: {folder_path}")
        return {}

    excel_files = list(folder_path.glob("*.xlsx"))
    if not excel_files:
        print(f" No Excel files found in folder: {folder_path}")
        return {}

    for file in excel_files:
        file_data = load_GDR_data_from_file(file)
        for code, titles in file_data.items():
            all_data.setdefault(code, []).extend(titles)

    # Preview summary
    print(f"\nLoaded GDR data from {len(excel_files)} files with {len(all_data)} unique classification codes.")
    for key, value in list(all_data.items())[:5]:  # Preview first 5 codes
        print(f"{key}: {value}")

    return all_data


def main(config=None):
    """Main function to load all Excel files in a folder."""
    folder_path = Path("C://Users//kripa//Research Project//ResearchLab//finalinput")  # Change to your folder path

    benennungen = load_GDR_data_from_folder(folder_path)

    if benennungen:
        print(" GDR data loaded successfully from folder.")
    else:
        print(" Failed to load any GDR data.")


if __name__ == "__main__":
    main()
